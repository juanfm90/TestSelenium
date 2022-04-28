
import unittest
import time
from DEPENDENCIAS import *
from datetime import date, timedelta
from selenium.webdriver import ActionChains
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from telnetlib import EC
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook


class BKI_RT_MAD_BIO_100(unittest.TestCase):

    def read_url(self):
        wb = load_workbook(filepath)
        sheet = wb.active
        A3 = sheet['A3']
        return A3

    def setUp(self):
        try:
            self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
            A3 = self.read_url()
            driver = self.driver
            driver.get(A3.value) #url value from excel
            self.driver.delete_all_cookies()
            self.driver.maximize_window()
            time.sleep(5)
        except Exception as e:
            print(f'Ha ocurrido un error: {e}')
        finally:
            print('webdriver: Ok')

    def read_origin(self):
        wb = load_workbook(filepath)
        sheet = wb.active
        A1 = sheet['A1']
        return A1

    def read_destination(self):
        wb = load_workbook(filepath)
        sheet = wb.active
        A2 = sheet['A2']
        return A2

    def read_email(self):
        wb = load_workbook(filepath)
        sheet = wb.active
        A6 = sheet['A6']
        return A6

    # def acceptCookies(self):
        # cookies = self.driver.find_element(By.XPATH, "//*[@id='onetrust-accept-btn-handler']")
        # cookies.click()
        # time.sleep(0.5)

    def depart_location(self):
        try:

            # cookies = self.driver.find_element(By.XPATH, "//*[@id='onetrust-accept-btn-handler']")
            # cookies.click()
            # time.sleep(0.5)

            depart_from = self.driver.find_element(By.XPATH, origen1_xpath)
            depart_from.click()
            depart_from.clear()
            A1 = self.read_origin()
            depart_from.send_keys(A1.value)
            depart_from.send_keys(Keys.ENTER)
            self.wait = WebDriverWait(self.driver, 2)
            time.sleep(0.5)
        except Exception as e:
            print(f'Ha ocurrido un error: {e}')

    def dest_location(self):
        A2 = self.read_destination()
        going_to = self.driver.find_element(By.XPATH, destino1_xpath)
        going_to.send_keys(A2.value)
        going_to.send_keys(Keys.ENTER)
        time.sleep(0.5)

    def depart_date(self):
        date_depart = self.driver.find_element(By.XPATH, fechaIda_xpath)
        depDate = date.today() + timedelta(days=30)
        print(depDate.strftime('%d%m%Y'))
        date_depart.send_keys(depDate.strftime('%d%m%Y'))
        time.sleep(0.5)

    def return_date(self):
        date_return = self.driver.find_element(By.XPATH, fechaVuelta_xpath)
        retDate = date.today() + timedelta(days=40)
        print(retDate.strftime('%d%m%Y'))
        date_return.send_keys(retDate.strftime('%d%m%Y'))
        date_return.click()
        time.sleep(0.5)

    def open_pass(self):
        passengers = self.driver.find_element(By.XPATH, pasajeros_dropdown)
        passengers.click()
        time.sleep(0.1)

    def add_passengers(self, num_adults, num_infants, num_babies):
        n_adults = self.driver.find_element(By.XPATH, add_adults)
        n_infants = self.driver.find_element(By.XPATH, add_children)
        n_babies = self.driver.find_element(By.XPATH, add_babies)
        for x in range(num_adults-1):
            n_adults.click()
        for x in range(num_infants):
            n_infants.click()
        for x in range(num_babies):
            n_babies.click()
        passengers = self.driver.find_element(By.XPATH, pasajeros_dropdown)
        time.sleep(0.1)
        passengers.click()

    def click_search(self):
        search = self.driver.find_element(By.XPATH, botonBuscar_xpath)
        time.sleep(1)
        search.click()

# -------Availability-------
    def ocultar_info(self):
        ocultar = WebDriverWait(self.driver, 50).until(
            EC.visibility_of_element_located((By.XPATH, ocultarInfo)))
        ocultar.click()
# -------------ANTIGUO------------
    # def results_ida(self, tarifa):
    #     vuelos_ida = self.driver.find_element(By.XPATH, div_ida)
    #     # vuelos_ida = self.driver.find_element(By.XPATH, "//*[@id='bbki-availability-trip-info']/div[2]")
    #
    #     ib = vuelos_ida.find_elements(By.TAG_NAME, "img")
    #     miIndex=0
    #     for x in ib:
    #         miIndex = miIndex+1
    #
    #         atr = x.get_attribute("ng-src")
    #         print(atr)
    #         if(atr == "/iberia-web-content/iconos/aerolineas/logo_pos_IB.svg"):
    #             miIndex=miIndex-1
    #             print(str(miIndex))
    #             break
    #
    #     print(str(miIndex))
    #     if(miIndex>=3):
    #         miIndex= miIndex/3+1
    #     else:
    #         miIndex=0
    #
    #     print(str(miIndex))
    #
    #     botones = vuelos_ida.find_elements(By.TAG_NAME, "button")
    #     select = int(miIndex)+2
    #     time.sleep(1)
    #
    #     try:
    #         botones[select].click()
    #         vuelos_ida = self.driver.find_element(By.XPATH, div_vuelos_ida_seleccionada)
    #
    #         # vuelos_ida = self.driver.find_element(By.XPATH, "//*[@id='selectedSlice-0']")
    #         # ib = vuelos_ida.find_elements(By.TAG_NAME, "input")
    #         time.sleep(0.2)
    #     except StaleElementReferenceException:
    #         vuelos_ida = self.driver.find_element(By.XPATH, div_vuelos_ida_seleccionada)
    #         # vuelos_ida = self.driver.find_element(By.XPATH, "//*[@id='selectedSlice-0']")
    #         # ib = vuelos_ida.find_elements(By.TAG_NAME, "input")
    #         pass
    #
    #     time.sleep(0.2)
    #
    #     radios_ida = vuelos_ida.find_elements(By.CLASS_NAME, radios_tarifa_ida)
    #     radios_ida[tarifa].click()

    # def results_vuelta(self, tarifa):
    #     # vuelos_vuelta = self.driver.find_element(By.XPATH, '//*[@id="bbki-availability-trip-info"]/div[2]')
    #     vuelos_vuelta = self.driver.find_element(By.XPATH, div_vuelta)
    #     ib = vuelos_vuelta.find_elements(By.TAG_NAME, "img")
    #
    #     miIndex = 0
    #     for x in ib:
    #         miIndex = miIndex + 1
    #
    #         atr = x.get_attribute("ng-src")
    #         print(atr)
    #         if (atr == "/iberia-web-content/iconos/aerolineas/logo_pos_IB.svg"):
    #             miIndex = miIndex - 1
    #             print(str(miIndex))
    #             break
    #
    #     print(str(miIndex))
    #     if (miIndex >= 3):
    #         miIndex = miIndex / 3 + 1
    #     else:
    #         miIndex = 0
    #
    #     print(str(miIndex))
    #
    #     botones = vuelos_vuelta.find_elements(By.TAG_NAME, "button")
    #     select = int(miIndex) + 2
    #     time.sleep(2)
    #
    #     try:
    #         botones[select].click()
    #         # vuelos_vuelta = self.driver.find_element(By.XPATH, '//*[@id="selectedSlice-1"]')
    #         vuelos_vuelta = self.driver.find_element(By.XPATH, div_vuelos_vuelta_seleccionada)
    #         # ib = vuelos_ida.find_elements(By.TAG_NAME, "input")
    #         time.sleep(2)
    #     except StaleElementReferenceException:
    #         # vuelos_vuelta = self.driver.find_element(By.XPATH, '//*[@id="selectedSlice-1"]')
    #         vuelos_vuelta = self.driver.find_element(By.XPATH, div_vuelos_vuelta_seleccionada)
    #         # ib = vuelos_ida.find_elements(By.TAG_NAME, "input")
    #         pass
    #
    #
    #     radios_vuelta = vuelos_vuelta.find_elements(By.CLASS_NAME, radios_tarifa_vuelta)
    #     radios_vuelta[tarifa].click()
    #     time.sleep(0.2)


    def results_ida(self, tarifa):
        URL = self.driver.current_url
        var = URL.split("/")[-1]
        print("Quadrigam: " + var)
        self.assertEqual("availability", var)

        vuelos_ida = self.driver.find_element(By.XPATH, div_ida)
        ib = vuelos_ida.find_elements(By.TAG_NAME, "img")
        miIndex = 0
        for x in ib:
            miIndex = miIndex + 1

            atr = x.get_attribute("ng-src")
            print(atr)
            if (atr == "/iberia-web-content/iconos/aerolineas/logo_pos_IB.svg"):
                miIndex = miIndex - 1
                print(str(miIndex))
                break
        if (miIndex >= 3):
            botones = vuelos_ida.find_elements(By.TAG_NAME, "button")
            print(len(botones))
            select = int(miIndex)  # /3 #+5
            print(str(select))
            int(select)
            time.sleep(1)
            botones[int(select) - 2].click() #-2 para turista, -1 para business
        else:
            botones = vuelos_ida.find_elements(By.TAG_NAME, "button")
            botones[2].click() #2 para turista, 3 para business
        print(str(miIndex))


        try:
            vuelos_ida = self.driver.find_element(By.XPATH, div_vuelos_ida_seleccionada)
            time.sleep(0.2)
        except StaleElementReferenceException:
            vuelos_ida = self.driver.find_element(By.XPATH, div_vuelos_ida_seleccionada)
            pass

        time.sleep(0.2)

        radios_ida = vuelos_ida.find_elements(By.CLASS_NAME, radios_tarifa_ida)
        radios_ida[tarifa].click()


    def results_vuelta(self, tarifa):
        time.sleep(2)
        vuelos_vuelta = self.driver.find_element(By.XPATH, div_vuelta)
        ib = vuelos_vuelta.find_elements(By.TAG_NAME, "img")

        actions = ActionChains(self.driver)
        actions.move_to_element(ib[0]).perform()


        miIndex = 0
        for x in ib:
            miIndex = miIndex + 1

            atr = x.get_attribute("ng-src")
            print(atr)
            if (atr == "/iberia-web-content/iconos/aerolineas/logo_pos_IB.svg"):
                miIndex = miIndex - 1
                print(str(miIndex))
                break


        if (miIndex >= 3):
            botones = vuelos_vuelta.find_elements(By.TAG_NAME, "button")
            # for element in vuelos_vuelta.find_elements(By.TAG_NAME, "button"):
            #     print(element.text)
            select = int(miIndex)  # /3 #+5
            print(str(select))
            int(select)
            time.sleep(1)
            botones[int(select) - 2].click() #-2 para turista, -1 para business
        else:
            botones = vuelos_vuelta.find_elements(By.TAG_NAME, "button")
            print(len(botones))
            # for element in vuelos_vuelta.find_elements(By.TAG_NAME, "button"):
            #     print(element.text)
            botones[2].click() #2 para turista, 3 para business

        print(str(miIndex))

        try:
            vuelos_vuelta = self.driver.find_element(By.XPATH, div_vuelos_vuelta_seleccionada)
            time.sleep(2)
        except StaleElementReferenceException:
            vuelos_vuelta = self.driver.find_element(By.XPATH, div_vuelos_vuelta_seleccionada)
            pass

        radios_vuelta = vuelos_vuelta.find_elements(By.CLASS_NAME, radios_tarifa_vuelta)
        radios_vuelta[tarifa].click()
        time.sleep(0.2)


    def continuar(self):
        btn_cont = self.driver.find_element(By.ID, botonContinuar)
        btn_cont.click()

# -------Passenger data-------
    def one_adult_passenger_data(self):
        A6 = self.read_email()

        nombre = WebDriverWait(self.driver, 50).until(
            EC.visibility_of_element_located((By.XPATH, nombrePasajero)))
        actions = ActionChains(self.driver)
        actions.move_to_element(nombre).perform()
        nombre.send_keys('AdultOne')

        apellidos = self.driver.find_element(By.XPATH, apellidoPasajero)
        actions = ActionChains(self.driver)
        actions.move_to_element(apellidos).perform()
        apellidos.send_keys('Test')

        email = self.driver.find_element(By.XPATH, contactoEmail)
        actions = ActionChains(self.driver)
        actions.move_to_element(email).perform()
        email.send_keys(A6.value)

        re_email = self.driver.find_element(By.XPATH, repetirEmail)
        actions = ActionChains(self.driver)
        actions.move_to_element(re_email).perform()
        re_email.send_keys(A6.value)

        tel = self.driver.find_element(By.XPATH, tel_xpath)
        actions = ActionChains(self.driver)
        actions.move_to_element(tel).perform()
        tel.send_keys('666666666')

        # no_publi = self.driver.find_element(By.XPATH, boletinNoticias)
        # actions = ActionChains(self.driver)
        # actions.move_to_element(no_publi).perform()
        # no_publi.click()

        continue_btn = self.driver.find_element(By.ID, botonContinuar)
        actions = ActionChains(self.driver)
        actions.move_to_element(continue_btn).perform()
        continue_btn.click()

# -------Ancillaries-------
    def close_flex(self):
        try:
            btn1 = WebDriverWait(self.driver, 50).until(
                EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[1]/form/header/div/button')))
            btn1.click()

        except:
            print("An exception occurred")
            pass

    def cont_btn(self):
        time.sleep(2)
        cont_bton = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '/html/body/main/div[2]/div[2]/ib-breakdown-ancillaries/div/div/section/div/div/div[2]/div/button')))
        # cont_bton = self.driver.find_element(By.XPATH, '/html/body/main/div[2]/div[2]/ib-breakdown-ancillaries/div/div/section/div/div/div[2]/div/button')
        actions = ActionChains(self.driver)
        actions.move_to_element(cont_bton).perform()
        cont_bton.click()

    def close_priority(self):
        try:
            prio_btn = WebDriverWait(self.driver, 15).until(
                EC.visibility_of_element_located((By.CLASS_NAME, 'ib-modals__button-action')))
            prio_btn.click()
        except:
            pass

    def close_seat(self):
        time.sleep(5)
        try:
            WebDriverWait(self.driver, 120).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="modalSeatMap"]/div/div/div/div[1]/div[2]')))
            seat_btn = WebDriverWait(self.driver, 15).until(
                EC.visibility_of_element_located((By.CLASS_NAME, 'ib-upgrade__button-close')))
            seat_btn.click()

        except:
            pass

# -------Payment-------
    def select_visa(self):
        time.sleep(5)
        # dropdown1 = WebDriverWait(self.driver, 20).until(
        #     EC.visibility_of_element_located((By.XPATH, '//*[@id="pmt-credit-card-tab-generic-credit-card-fields-1"]/div[2]/div[1]')))
        # dropdown1.click()
        #
        # cards = dropdown1.find_elements(By.CLASS_NAME, 'ib-select__list-txt')
        # cards[0].click()

        time.sleep(1)
        WebDriverWait(self.driver, 120).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="ibdc-number-frame"]')))

        iframe2 = self.driver.find_element(By.XPATH, '//*[@id="ibdc-number-frame"]')
        self.driver.switch_to.frame(iframe2)


        num_card = WebDriverWait(self.driver, 5).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="number"]')))
        actions = ActionChains(self.driver)
        actions.move_to_element(num_card).perform()
        num_card.send_keys('4012999999999999')

        self.driver.switch_to.parent_frame()

        nom_card = WebDriverWait(self.driver, 5).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="name"]')))
        actions = ActionChains(self.driver)
        actions.move_to_element(nom_card).perform()
        nom_card.send_keys('HANSOLO')

        surname_card = WebDriverWait(self.driver, 5).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="surnames"]')))
        actions = ActionChains(self.driver)
        actions.move_to_element(surname_card).perform()
        surname_card.send_keys('HANSOLO')

        open_month = self.driver.find_element(By.XPATH, '//*[@id="EXPIRY_DATE"]/div[1]/div[1]/span')
        open_month.click()
        # exp_month = self.driver.find_element(By.XPATH, '//*[@id="ui-select-choices-row-6-2"]/span')
        exp_month = self.driver.find_element(By.XPATH, '//*[@id="ui-select-choices-row-3-2"]')
        exp_month.click()
        open_year = self.driver.find_element(By.XPATH, '//*[@id="EXPIRY_DATE"]/div[2]/div[1]/span')
        open_year.click()
        # exp_year = self.driver.find_element(By.XPATH, '//*[@id="ui-select-choices-row-7-2"]/span/span')
        exp_year = self.driver.find_element(By.XPATH, '//*[@id="ui-select-choices-row-4-4"]/span/span')
        exp_year.click()

        iframe3 = self.driver.find_element(By.XPATH, '//*[@id="ibdc-cvv-frame"]')
        self.driver.switch_to.frame(iframe3)

        cvv = self.driver.find_element(By.XPATH, '//*[@id="cvv"]')
        cvv.send_keys('123')

        self.driver.switch_to.parent_frame()

# -----Write data----
    def write_pnr(self):
        get_text = self.driver.find_element(By.XPATH, '//*[@id="ui-id-2"]').text
        wb = load_workbook(filepath)
        sheet = wb.active
        sheet["A8"] = get_text
        wb.save("Variables.xlsx")

    def test_1_BKI_RT(self):
        # self.acceptCookies()
        self.write_pnr()

        self.depart_location()
        self.dest_location()
        self.depart_date()
        self.return_date()
        self.open_pass()
        self.add_passengers(1,0,0)
        self.click_search()

        self.ocultar_info()
        self.results_ida(0)
        self.results_vuelta(0)
        self.continuar()

        self.one_adult_passenger_data()
        # self.close_priority()
        # self.close_flex()
        self.close_seat()
        # self.close_flex()
        self.cont_btn()

        self.select_visa()
        # time.sleep(1000)

if __name__ == '__main__':
    unittest.main()
