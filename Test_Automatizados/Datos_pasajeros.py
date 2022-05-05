import time
from DEPENDENCIAS import *
from datetime import date, timedelta
from selenium.webdriver import ActionChains
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from telnetlib import EC
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

class Datos_pasajeros():

    def read_email(self):
        wb = load_workbook(filepath)
        sheet = wb.active
        A6 = sheet['A6']
        return A6

    def read_tr_cantidad_primeros(self):
        wb = load_workbook(filepath)
        sheet = wb.active
        H2 = sheet['H2']
        return H2

    def read_tr_email_primeros(self):
        wb = load_workbook(filepath)
        sheet = wb.active
        H3 = sheet['H3']
        return H3

    def read_tr_cantidad_terceros(self):
        wb = load_workbook(filepath)
        sheet = wb.active
        F2 = sheet['F2']
        return F2

    def read_tr_email_remitente(self):
        wb = load_workbook(filepath)
        sheet = wb.active
        F3 = sheet['F3']
        return F3

    def read_tr_email_receptor(self):
        wb = load_workbook(filepath)
        sheet = wb.active
        F4 = sheet['F4']
        return F4

    def read_tr_nombre_remitente(self):
        wb = load_workbook(filepath)
        sheet = wb.active
        F5 = sheet['F5']
        return F5

    def read_tr_nombre_receptor(self):
        wb = load_workbook(filepath)
        sheet = wb.active
        F6 = sheet['F6']
        return F6

    def read_tr_mensaje(self):
        wb = load_workbook(filepath)
        sheet = wb.active
        F7 = sheet['F7']
        return F7

    def read_tr_impresion_check(self):
        wb = load_workbook(filepath)
        sheet = wb.active
        F8 = sheet['F8']
        return F8

    def read_email_impresion(self):
        wb = load_workbook(filepath)
        sheet = wb.active
        F9 = sheet['F9']
        return F9

    def read_bonos_campaña_tipo(self):
        wb = load_workbook(filepath)
        sheet = wb.active
        F13 = sheet['F13']
        return F13

    def read_bonos_campaña_cantidad_total(self):
        wb = load_workbook(filepath)
        sheet = wb.active
        F14 = sheet['F14']
        return F14

    def read_bonos_campaña_cantidad_porcentual(self):
        wb = load_workbook(filepath)
        sheet = wb.active
        F15 = sheet['F15']
        return F15

    def read_bonos_campaña_cantidad_marketing(self):
        wb = load_workbook(filepath)
        sheet = wb.active
        F16 = sheet['F16']
        return F16

    def read_bonos_campaña_numero_usos(self):
        wb = load_workbook(filepath)
        sheet = wb.active
        F17 = sheet['F17']
        return F17

    def read_bonos_edicion_cantidad_porcentual(self):
        wb = load_workbook(filepath)
        sheet = wb.active
        F20 = sheet['F20']
        return F20
    def read_bonos_edicion_cantidad_total(self):
        wb = load_workbook(filepath)
        sheet = wb.active
        F21 = sheet['F21']
        return F21

    def read_bonos_edicion_cantidad_marketing(self):
        wb = load_workbook(filepath)
        sheet = wb.active
        F22 = sheet['F22']
        return F22

    def read_bonos_edicion_numero_usos(self):
        wb = load_workbook(filepath)
        sheet = wb.active
        F23 = sheet['F23']
        return F23


    def cont_vuelo(driver):
        try:
            WebDriverWait(driver, 15).until(
                EC.element_to_be_clickable((By.XPATH, '/html/body/main/div[2]/section/section/div/div[1]/div/div/div/div[2]')))
            cont = driver.find_element(By.XPATH, '/html/body/main/div[2]/section/section/div/div[1]/div/div/div/div[2]')
            time.sleep(3)
            cont.click()

        except:
            print('Botón no mostrado')

    def one_adult_passenger_data(driver, self=None):
        A6 = Datos_pasajeros.read_email(self)
        WebDriverWait(driver, 50).until(
            EC.visibility_of_element_located((By.XPATH, nombrePasajero)))

        nombre = WebDriverWait(driver, 50).until(
            EC.visibility_of_element_located((By.XPATH, nombrePasajero)))
        actions = ActionChains(driver)
        actions.move_to_element(nombre).perform()
        nombre.send_keys('AdultOne')

        apellidos = driver.find_element(By.XPATH, apellidoPasajero)
        actions = ActionChains(driver)
        actions.move_to_element(apellidos).perform()
        apellidos.send_keys('Test')

        email = driver.find_element(By.XPATH, contactoEmail)
        actions = ActionChains(driver)
        actions.move_to_element(email).perform()
        email.send_keys(A6.value)

        re_email = driver.find_element(By.XPATH, repetirEmail)
        actions = ActionChains(driver)
        actions.move_to_element(re_email).perform()
        re_email.send_keys(A6.value)

        tel = driver.find_element(By.XPATH, tel_xpath)
        actions = ActionChains(driver)
        actions.move_to_element(tel).perform()
        tel.send_keys('666666666')

        no_publi = driver.find_element(By.XPATH, boletinNoticias)
        actions = ActionChains(driver)
        actions.move_to_element(no_publi).perform()
        no_publi.click()

        continue_btn = driver.find_element(By.ID, botonContinuar)
        actions = ActionChains(driver)
        actions.move_to_element(continue_btn).perform()
        continue_btn.click()

    def two_adults_passenger_data(driver, self=None):
        A6 = Datos_pasajeros.read_email(self)
        WebDriverWait(driver, 50).until(
            EC.visibility_of_element_located((By.XPATH, nombrePasajero)))

        nombre = WebDriverWait(driver, 50).until(
            EC.visibility_of_element_located((By.XPATH, nombrePasajero)))
        actions = ActionChains(driver)
        actions.move_to_element(nombre).perform()
        nombre.send_keys('AdultOne')

        apellidos = driver.find_element(By.XPATH, apellidoPasajero)
        actions = ActionChains(driver)
        actions.move_to_element(apellidos).perform()
        apellidos.send_keys('Test')

        nombre2 = WebDriverWait(driver, 50).until(
            EC.visibility_of_element_located((By.XPATH, nombre2adulto)))
        actions = ActionChains(driver)
        actions.move_to_element(nombre).perform()
        nombre2.send_keys('AdultTwo')

        apellidos2 = driver.find_element(By.XPATH, apellido2adulto)
        actions = ActionChains(driver)
        actions.move_to_element(apellidos).perform()
        apellidos2.send_keys('Test')

        email = driver.find_element(By.XPATH, contactoEmail)
        actions = ActionChains(driver)
        actions.move_to_element(email).perform()
        email.send_keys(A6.value)

        re_email = driver.find_element(By.XPATH, repetirEmail)
        actions = ActionChains(driver)
        actions.move_to_element(re_email).perform()
        re_email.send_keys(A6.value)

        tel = driver.find_element(By.XPATH, tel_xpath)
        actions = ActionChains(driver)
        actions.move_to_element(tel).perform()
        tel.send_keys('666666666')

        no_publi = driver.find_element(By.XPATH, boletinNoticias)
        actions = ActionChains(driver)
        actions.move_to_element(no_publi).perform()
        no_publi.click()

        continue_btn = driver.find_element(By.ID, botonContinuar)
        actions = ActionChains(driver)
        actions.move_to_element(continue_btn).perform()
        continue_btn.click()

    def oneadult_onechild_onebaby_passenger(driver, self=None):
        A6 = Datos_pasajeros.read_email(self)

        # -----------------NOMBRE ADULTO------------------
        WebDriverWait(driver, 50).until(
            EC.visibility_of_element_located((By.XPATH, nombrePasajero)))
        nombre = WebDriverWait(driver, 50).until(
            EC.visibility_of_element_located((By.XPATH, nombrePasajero)))
        actions = ActionChains(driver)
        actions.move_to_element(nombre).perform()
        nombre.send_keys('AdultOne')

        apellidos = driver.find_element(By.XPATH, apellidoPasajero)
        actions = ActionChains(driver)
        actions.move_to_element(apellidos).perform()
        apellidos.send_keys('Test')

        # -----------------NOMBRE NIÑO-----------------------
        nombre2 = WebDriverWait(driver, 50).until(
            EC.visibility_of_element_located((By.XPATH, nombre2adulto)))
        actions = ActionChains(driver)
        actions.move_to_element(nombre2).perform()
        nombre2.send_keys('ChildOne')

        apellidos2 = driver.find_element(By.XPATH, apellido2adulto)
        actions = ActionChains(driver)
        actions.move_to_element(apellidos2).perform()
        apellidos2.send_keys('Test')

        diaNac1 = driver.find_element(By.XPATH, '//*[@id="PASSENGER_BIRTH_DATE"]/div[1]/div[1]')
        actions = ActionChains(driver)
        actions.move_to_element(diaNac1).perform()
        diaNac1.click()
        birthDay = driver.find_element(By.XPATH, '//*[@id="ui-select-choices-row-13-5"]/span/span')
        # //*[@id="ui-select-choices-row-4-0"]/span"]
        birthDay.click()

        mesNac1 = driver.find_element(By.XPATH, '//*[@id="PASSENGER_BIRTH_DATE"]/div[2]/div[1]')
        mesNac1.click()
        birthMonth = driver.find_element(By.XPATH, '//*[@id="ui-select-choices-row-14-0"]/span/span')
        birthMonth.click()

        anoNac1 = driver.find_element(By.XPATH, '//*[@id="PASSENGER_BIRTH_DATE"]/div[3]/div[1]')
        anoNac1.click()
        birthYear = driver.find_element(By.XPATH, '//*[@id="ui-select-choices-row-15-0"]/span/span')
        birthYear.click()

        # -----------------NOMBRE BEBÉ-------------------
        nombre3 = driver.find_element(By.XPATH, '//*[@id="name_2"]')
        actions = ActionChains(driver)
        actions.move_to_element(nombre3).perform()
        nombre3.send_keys('BabyOne')

        apellidos3 = driver.find_element(By.XPATH, '//*[@id="first_surname_2"]')
        actions = ActionChains(driver)
        actions.move_to_element(apellidos3).perform()
        apellidos3.send_keys('BabyOne')

        diaNac1 = driver.find_element(By.XPATH, '(//*[@id="PASSENGER_BIRTH_DATE"]/div[1]/div[1])[2]')
        actions = ActionChains(driver)
        actions.move_to_element(diaNac1).perform()
        diaNac1.click()
        birthDay = driver.find_element(By.XPATH, '//*[@id="ui-select-choices-row-16-0"]/span/span')
        birthDay.click()

        mesNac1 = driver.find_element(By.XPATH, '(//*[@id="PASSENGER_BIRTH_DATE"]/div[2]/div[1])[2]')
        mesNac1.click()
        birthMonth = driver.find_element(By.XPATH, '//*[@id="ui-select-choices-row-17-0"]')
        birthMonth.click()

        anoNac1 = driver.find_element(By.XPATH, '(//*[@id="PASSENGER_BIRTH_DATE"]/div[3]/div[1])[2]')
        anoNac1.click()
        birthYear = driver.find_element(By.XPATH, '//*[@id="ui-select-choices-row-18-0"]')
        birthYear.click()

        email = driver.find_element(By.XPATH, contactoEmail)
        actions = ActionChains(driver)
        actions.move_to_element(email).perform()
        email.clear()
        email.send_keys(A6.value)

        re_email = driver.find_element(By.XPATH, repetirEmail)
        actions = ActionChains(driver)
        actions.move_to_element(re_email).perform()
        re_email.clear()
        re_email.send_keys(A6.value)

        tel = driver.find_element(By.XPATH, tel_xpath)
        actions = ActionChains(driver)
        actions.move_to_element(tel).perform()
        tel.clear()
        tel.send_keys('666666666')

        # no_publi = self.driver.find_element(By.XPATH, '//*[@id="bbki-passenger-info-passengers-contact-form"]/fieldset/fieldset/div/div/div/div[2]/div/div[5]/label')
        # actions = ActionChains(self.driver)
        # actions.move_to_element(no_publi).perform()
        # no_publi.click()

        # URL = self.driver.current_url
        # var = URL.split("/")[-1]
        # print("Quadrigam: " + var)
        # self.assertEqual("ibairp", var)

        continue_btn = driver.find_element(By.ID, botonContinuar)
        actions = ActionChains(driver)
        actions.move_to_element(continue_btn).perform()
        continue_btn.click()

        try:
            no_salvar = self.driver.find_element(By.XPATH, '//*[@id="bbki-generic-complete-cancel-link"]')
            no_salvar.click()
        except:
            print('No se ha podido realizar')

    def oneadult_logado_onechild_onebaby_passenger(driver, self=None):
        A6 = Datos_pasajeros.read_email(self)

        # -----------------NOMBRE ADULTO------------------
        # WebDriverWait(driver, 50).until(
        #     EC.visibility_of_element_located((By.XPATH, nombrePasajero)))
        # nombre = WebDriverWait(driver, 50).until(
        #     EC.visibility_of_element_located((By.XPATH, nombrePasajero)))
        # actions = ActionChains(driver)
        # actions.move_to_element(nombre).perform()
        # nombre.send_keys('AdultOne')
        #
        # apellidos = self.driver.find_element(By.XPATH, apellidoPasajero)
        # actions = ActionChains(driver)
        # actions.move_to_element(apellidos).perform()
        # apellidos.send_keys('Test')

        # -----------------NOMBRE NIÑO-----------------------
        nombre2 = WebDriverWait(driver, 50).until(
            EC.visibility_of_element_located((By.XPATH, nombre2adulto)))
        actions = ActionChains(driver)
        actions.move_to_element(nombre2).perform()
        nombre2.send_keys('ChildOne')

        apellidos2 = driver.find_element(By.XPATH, apellido2adulto)
        actions = ActionChains(driver)
        actions.move_to_element(apellidos2).perform()
        apellidos2.send_keys('Test')

        diaNac1 = driver.find_element(By.XPATH, '//*[@id="PASSENGER_BIRTH_DATE"]/div[1]/div[1]')
        actions = ActionChains(driver)
        actions.move_to_element(diaNac1).perform()
        diaNac1.click()
        birthDay = driver.find_element(By.XPATH, '//*[@id="ui-select-choices-row-4-0"]/span')
        # //*[@id="ui-select-choices-row-4-0"]/span"]
        birthDay.click()

        mesNac1 = driver.find_element(By.XPATH, '//*[@id="PASSENGER_BIRTH_DATE"]/div[2]/div[1]')
        mesNac1.click()
        birthMonth = driver.find_element(By.XPATH, '//*[@id="ui-select-choices-row-5-0"]/span')
        birthMonth.click()

        anoNac1 = driver.find_element(By.XPATH, '//*[@id="PASSENGER_BIRTH_DATE"]/div[3]/div[1]')
        anoNac1.click()
        birthYear = driver.find_element(By.XPATH, '//*[@id="ui-select-choices-row-6-0"]/span')
        birthYear.click()

        # -----------------NOMBRE BEBÉ-------------------
        nombre3 = driver.find_element(By.XPATH, '//*[@id="name_2"]')
        actions = ActionChains(driver)
        actions.move_to_element(nombre3).perform()
        nombre3.send_keys('BabyOne')

        apellidos3 = driver.find_element(By.XPATH, '//*[@id="first_surname_2"]')
        actions = ActionChains(driver)
        actions.move_to_element(apellidos3).perform()
        apellidos3.send_keys('BabyOne')

        diaNac1 = driver.find_element(By.XPATH, '(//*[@id="PASSENGER_BIRTH_DATE"]/div[1]/div[1])[2]')
        actions = ActionChains(driver)
        actions.move_to_element(diaNac1).perform()
        diaNac1.click()
        birthDay = driver.find_element(By.XPATH, '//*[@id="ui-select-choices-row-7-0"]/span/span')
        birthDay.click()

        mesNac1 = driver.find_element(By.XPATH, '(//*[@id="PASSENGER_BIRTH_DATE"]/div[2]/div[1])[2]')
        mesNac1.click()
        birthMonth = driver.find_element(By.XPATH, '//*[@id="ui-select-choices-row-8-0"]/span/span')
        birthMonth.click()

        anoNac1 = driver.find_element(By.XPATH, '(//*[@id="PASSENGER_BIRTH_DATE"]/div[3]/div[1])[2]')
        anoNac1.click()
        birthYear = driver.find_element(By.XPATH, '//*[@id="ui-select-choices-row-9-0"]/span/span')
        birthYear.click()

        email = driver.find_element(By.XPATH, contactoEmail)
        actions = ActionChains(driver)
        actions.move_to_element(email).perform()
        email.clear()
        email.send_keys(A6.value)

        re_email = driver.find_element(By.XPATH, repetirEmail)
        actions = ActionChains(driver)
        actions.move_to_element(re_email).perform()
        re_email.clear()
        re_email.send_keys(A6.value)

        tel = driver.find_element(By.XPATH, tel_xpath)
        actions = ActionChains(driver)
        actions.move_to_element(tel).perform()
        tel.clear()
        tel.send_keys('666666666')

        # no_publi = self.driver.find_element(By.XPATH, '//*[@id="bbki-passenger-info-passengers-contact-form"]/fieldset/fieldset/div/div/div/div[2]/div/div[5]/label')
        # actions = ActionChains(self.driver)
        # actions.move_to_element(no_publi).perform()
        # no_publi.click()

        # URL = self.driver.current_url
        # var = URL.split("/")[-1]
        # print("Quadrigam: " + var)
        # self.assertEqual("ibairp", var)

        continue_btn = driver.find_element(By.ID, botonContinuar)
        actions = ActionChains(driver)
        actions.move_to_element(continue_btn).perform()
        continue_btn.click()
        time.sleep(1)

        try:
            no_salvar = driver.find_element(By.XPATH, '//*[@id="bbki-generic-complete-cancel-link"]')
            no_salvar.click()
        except:
            print('No se ha podido realizar')

    def one_adult_FF(driver, self=None):
        A6 = Datos_pasajeros.read_email(self)
        WebDriverWait(driver, 50).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="name_0"]')))

        email = driver.find_element(By.XPATH, contactoEmail)
        email.location_once_scrolled_into_view
        actions = ActionChains(driver)
        actions.move_to_element(email).perform()
        email.clear()
        email.send_keys(A6.value)

        re_email = driver.find_element(By.XPATH, repetirEmail)
        actions = ActionChains(driver)
        actions.move_to_element(re_email).perform()
        re_email.clear()
        re_email.send_keys(A6.value)

        tel = driver.find_element(By.XPATH, tel_xpath)
        actions = ActionChains(driver)
        actions.move_to_element(tel).perform()
        tel.clear()
        tel.send_keys('666666666')

        # no_publi = driver.find_element(By.XPATH, boletinNoticias)
        # actions = ActionChains(driver)
        # actions.move_to_element(no_publi).perform()
        # no_publi.click()

        continue_btn = driver.find_element(By.ID, botonContinuar)
        actions = ActionChains(driver)
        actions.move_to_element(continue_btn).perform()
        continue_btn.click()

        try:
            no_salvar = driver.find_element(By.XPATH, '//*[@id="bbki-generic-complete-cancel-link"]')
            no_salvar.click()
        except:
            print('No se ha podido realizar')